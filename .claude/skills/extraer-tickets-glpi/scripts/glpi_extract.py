#!/usr/bin/env python3
"""
Extrae tickets de GLPI via REST API y genera reporte en formato Markdown.

Uso:
    python3 scripts/glpi_extract.py ENTIDAD FECHA_INICIO FECHA_FIN [OUTPUT_PATH]

Ejemplo:
    python3 scripts/glpi_extract.py IDARTES 2026-02-01 2026-02-28 output.md

Variables de entorno requeridas:
    GLPI_{ENTIDAD}_URL        - URL base de la API (ej: https://glpi.idartes.gov.co/apirest.php)
    GLPI_{ENTIDAD}_APP_TOKEN  - App Token generado por el admin GLPI
    GLPI_{ENTIDAD}_USER_TOKEN - User Token del perfil del usuario
"""

import argparse
import json
import os
import sys
from datetime import datetime
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen


def get_env(entidad: str) -> dict:
    """Lee credenciales de variables de entorno."""
    prefix = f"GLPI_{entidad.upper()}"
    url = os.environ.get(f"{prefix}_URL")
    app_token = os.environ.get(f"{prefix}_APP_TOKEN")
    user_token = os.environ.get(f"{prefix}_USER_TOKEN")

    missing = []
    if not url:
        missing.append(f"{prefix}_URL")
    if not app_token:
        missing.append(f"{prefix}_APP_TOKEN")
    if not user_token:
        missing.append(f"{prefix}_USER_TOKEN")

    if missing:
        print(f"ERROR: Variables de entorno faltantes: {', '.join(missing)}", file=sys.stderr)
        print(f"Configurar en archivo .env o exportar manualmente.", file=sys.stderr)
        sys.exit(1)

    return {"url": url, "app_token": app_token, "user_token": user_token}


def api_request(url: str, headers: dict, method: str = "GET") -> dict:
    """Realiza una peticion HTTP a la API de GLPI."""
    req = Request(url, headers=headers, method=method)
    try:
        with urlopen(req, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        print(f"ERROR HTTP {e.code}: {body}", file=sys.stderr)
        raise
    except URLError as e:
        print(f"ERROR de conexion: {e.reason}", file=sys.stderr)
        raise


def init_session(config: dict) -> str:
    """Inicia sesion en GLPI y retorna el session_token."""
    headers = {
        "App-Token": config["app_token"],
        "Authorization": f"user_token {config['user_token']}",
        "Content-Type": "application/json",
    }
    data = api_request(f"{config['url']}/initSession", headers)
    return data["session_token"]


def kill_session(config: dict, session_token: str) -> None:
    """Cierra la sesion de GLPI."""
    headers = {
        "App-Token": config["app_token"],
        "Session-Token": session_token,
        "Content-Type": "application/json",
    }
    try:
        api_request(f"{config['url']}/killSession", headers)
    except Exception:
        pass  # No fallar si kill_session falla


def search_tickets(config: dict, session_token: str, fecha_inicio: str, fecha_fin: str) -> list:
    """Busca tickets por rango de fechas usando la API de busqueda."""
    headers = {
        "App-Token": config["app_token"],
        "Session-Token": session_token,
        "Content-Type": "application/json",
    }

    # Criterios de busqueda GLPI:
    # - Campo 15 = date (fecha apertura)
    # - Campo 16 = closedate (fecha cierre)
    # - Campo 12 = status
    # - Campo 1 = name (titulo)
    # - Campo 2 = id
    # - Campo 7 = itilcategories_id (categoria)
    # - Campo 18 = time_to_resolve (tiempo resolucion)
    # Usamos fecha de apertura (campo 15) en el rango

    from urllib.parse import quote

    params = (
        "?criteria[0][field]=15"
        "&criteria[0][searchtype]=morethan"
        f"&criteria[0][value]={fecha_inicio}"
        "&criteria[1][link]=AND"
        "&criteria[1][field]=15"
        "&criteria[1][searchtype]=lessthan"
        f"&criteria[1][value]={fecha_fin}"
        "&forcedisplay[0]=2"   # ID
        "&forcedisplay[1]=1"   # Titulo
        "&forcedisplay[2]=15"  # Fecha apertura
        "&forcedisplay[3]=16"  # Fecha cierre
        "&forcedisplay[4]=12"  # Estado
        "&forcedisplay[5]=7"   # Categoria
        "&forcedisplay[6]=18"  # Tiempo resolucion
        "&range=0-200"
        "&sort=15"
        "&order=ASC"
    )

    url = f"{config['url']}/search/Ticket{params}"
    data = api_request(url, headers)

    if not data or "data" not in data:
        return []

    return data["data"]


def status_label(status_id) -> str:
    """Convierte ID de estado GLPI a texto."""
    status_map = {
        1: "Nuevo",
        2: "En curso (asignado)",
        3: "En curso (planificado)",
        4: "En espera",
        5: "Resuelto",
        6: "Cerrado",
    }
    try:
        return status_map.get(int(status_id), str(status_id))
    except (ValueError, TypeError):
        return str(status_id)


def generate_markdown(tickets: list, entidad: str, fecha_inicio: str, fecha_fin: str) -> str:
    """Genera el reporte Markdown a partir de los tickets."""
    total = len(tickets)

    # Contar por estado
    estados = {}
    for t in tickets:
        estado = status_label(t.get("12", ""))
        estados[estado] = estados.get(estado, 0) + 1

    # Contar cerrados/resueltos
    cerrados = sum(1 for t in tickets if int(t.get("12", 0)) in (5, 6))
    pct_cumplimiento = round((cerrados / total) * 100, 1) if total > 0 else 0

    lines = [
        f"# Reporte de Soportes GLPI - {entidad}",
        f"## Periodo: {fecha_inicio} a {fecha_fin}",
        f"## Fuente: API GLPI REST",
        "",
        "---",
        "",
        "## Resumen General",
        "",
        f"Durante el periodo reportado se atendieron **{total}** solicitudes a traves "
        f"de la mesa de servicios GLPI. Se logro un cumplimiento del **{pct_cumplimiento}%** "
        f"en la resolucion de tickets.",
        "",
        "## Estadisticas",
        "",
        f"- **Total tickets:** {total}",
    ]

    for estado, count in sorted(estados.items()):
        lines.append(f"- **{estado}:** {count}")

    lines.extend([
        "",
        "## Tabla de Tickets",
        "",
        "| # | ID | Fecha Apertura | Fecha Cierre | Estado | Titulo |",
        "|---|-----|----------------|--------------|--------|--------|",
    ])

    for i, t in enumerate(tickets, 1):
        tid = t.get("2", "")
        titulo = t.get("1", "").replace("|", "\\|")
        fecha_apertura = t.get("15", "")[:10] if t.get("15") else ""
        fecha_cierre = t.get("16", "")[:10] if t.get("16") else "-"
        estado = status_label(t.get("12", ""))

        lines.append(f"| {i} | {tid} | {fecha_apertura} | {fecha_cierre} | {estado} | {titulo} |")

    lines.extend([
        "",
        "---",
        "",
        f"*Reporte generado automaticamente via GLPI REST API el {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
    ])

    return "\n".join(lines)


def generate_xlsx(tickets, entidad, fecha_inicio, fecha_fin, xlsx_path):
    """Genera un Excel con la tabla de tickets GLPI."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets GLPI"

    # Estadisticas en header
    title_font = Font(bold=True, size=14, color="003366")
    ws["A1"] = f"REPORTE DE TICKETS GLPI - {entidad}"
    ws["A1"].font = title_font
    ws["A2"] = f"Periodo: {fecha_inicio} a {fecha_fin}"
    ws["A2"].font = Font(size=10, color="666666")

    total = len(tickets)
    cerrados = sum(1 for t in tickets if int(t.get("12", 0)) in (5, 6))
    pct = round((cerrados / total) * 100, 1) if total > 0 else 0

    ws["A3"] = f"Total tickets: {total}  |  Cerrados/Resueltos: {cerrados}  |  Cumplimiento: {pct}%"
    ws["A3"].font = Font(size=10, bold=True)

    # Tabla de datos
    header_row = 5
    headers = ["#", "ID", "Fecha Apertura", "Fecha Cierre", "Estado", "Titulo", "Categoria"]
    col_widths = [6, 10, 18, 18, 22, 50, 25]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for i, t in enumerate(tickets, 1):
        row = header_row + i
        fill_color = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        values = [
            i,
            t.get("2", ""),
            (t.get("15", "") or "")[:10],
            (t.get("16", "") or "-")[:10] if t.get("16") else "-",
            status_label(t.get("12", "")),
            (t.get("1", "") or ""),
            (t.get("7", "") or ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-filtro y freeze
    last_row = header_row + len(tickets)
    ws.auto_filter.ref = f"A{header_row}:G{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb.save(xlsx_path)
    print(f"Excel generado: {xlsx_path} ({len(tickets)} tickets)", file=sys.stderr)


def parse_args():
    """Parsea argumentos con compatibilidad posicional."""
    parser = argparse.ArgumentParser(
        description="Extrae tickets de GLPI via REST API",
        usage="python3 scripts/glpi_extract.py ENTIDAD FECHA_INICIO FECHA_FIN [OUTPUT_PATH] [--output-xlsx PATH]",
    )
    parser.add_argument("entidad", help="Nombre de la entidad (IDARTES, SDMUJER)")
    parser.add_argument("fecha_inicio", help="Fecha inicio (YYYY-MM-DD)")
    parser.add_argument("fecha_fin", help="Fecha fin (YYYY-MM-DD)")
    parser.add_argument("output_path", nargs="?", default=None, help="Ruta de salida del archivo .md")
    parser.add_argument("--output-xlsx", dest="output_xlsx", default=None, help="Ruta de salida del archivo Excel (.xlsx)")
    return parser.parse_args()


def main():
    args = parse_args()

    entidad = args.entidad.upper()
    fecha_inicio = args.fecha_inicio
    fecha_fin = args.fecha_fin
    output_path = args.output_path
    xlsx_path = args.output_xlsx

    # Validar formato de fechas
    for fecha in [fecha_inicio, fecha_fin]:
        try:
            datetime.strptime(fecha, "%Y-%m-%d")
        except ValueError:
            print(f"ERROR: Formato de fecha invalido: {fecha}. Usar YYYY-MM-DD", file=sys.stderr)
            sys.exit(1)

    config = get_env(entidad)
    session_token = None

    try:
        print(f"Conectando a GLPI ({entidad})...", file=sys.stderr)
        session_token = init_session(config)
        print(f"Sesion iniciada. Buscando tickets del {fecha_inicio} al {fecha_fin}...", file=sys.stderr)

        tickets = search_tickets(config, session_token, fecha_inicio, fecha_fin)
        print(f"Tickets encontrados: {len(tickets)}", file=sys.stderr)

        markdown = generate_markdown(tickets, entidad, fecha_inicio, fecha_fin)

        if output_path:
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(markdown)
            print(f"Reporte guardado en: {output_path}", file=sys.stderr)
        else:
            print(markdown)

        if xlsx_path and tickets:
            generate_xlsx(tickets, entidad, fecha_inicio, fecha_fin, xlsx_path)

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        if session_token:
            kill_session(config, session_token)
            print("Sesion GLPI cerrada.", file=sys.stderr)


if __name__ == "__main__":
    main()
