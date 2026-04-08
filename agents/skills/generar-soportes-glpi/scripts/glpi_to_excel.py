#!/usr/bin/env python3
"""
Genera un Excel con tickets GLPI a partir de datos JSON.

Uso:
    python3 scripts/glpi_to_excel.py <json_input> <xlsx_output>

Input JSON:
    {
        "entidad": "IDARTES",
        "periodo_inicio": "2026-02-01",
        "periodo_fin": "2026-02-28",
        "total": 11,
        "tickets": [
            {
                "id": "12345",
                "titulo": "Error en modulo Planeacion",
                "tipo": "Incidencia",
                "estado": "Cerrado",
                "fecha_apertura": "2026-02-03",
                "fecha_cierre": "2026-02-05",
                "prioridad": "Media",
                "categoria": "Software",
                "solicitante": "usuario@idartes.gov.co",
                "solucion": "Se ajusto la consulta SQL..."
            }
        ]
    }
"""

import json
import os
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_header_style():
    """Retorna estilos para el header de la tabla."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def create_cell_style(row_idx):
    """Retorna estilo para celdas de datos con sombreado alternado."""
    fill_color = "F5F5F5" if row_idx % 2 == 0 else "FFFFFF"
    return {
        "fill": PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "alignment": Alignment(vertical="center", wrap_text=True),
    }


def generate_excel(data, output_path):
    """Genera el Excel a partir de los datos JSON."""
    entidad = data.get("entidad", "")
    periodo_inicio = data.get("periodo_inicio", "")
    periodo_fin = data.get("periodo_fin", "")
    total = data.get("total", 0)
    tickets = data.get("tickets", [])

    wb = Workbook()

    # --- Hoja 1: Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    title_font = Font(bold=True, size=14, color="003366")
    subtitle_font = Font(bold=True, size=11, color="333333")
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)

    ws_resumen["A1"] = f"REPORTE DE TICKETS GLPI - {entidad}"
    ws_resumen["A1"].font = title_font
    ws_resumen["A2"] = f"Período: {periodo_inicio} a {periodo_fin}"
    ws_resumen["A2"].font = Font(size=10, color="666666")

    row = 4
    ws_resumen.cell(row=row, column=1, value="Estadísticas Generales").font = subtitle_font
    row += 1

    # Contar cerrados
    cerrados = sum(1 for t in tickets if t.get("estado", "").lower() in ("cerrado", "resuelto", "closed", "solved"))
    stats = [
        ("Total tickets:", total),
        ("Cerrados/Resueltos:", cerrados),
        ("% Resueltos:", f"{round((cerrados / total) * 100, 1)}%" if total > 0 else "0%"),
    ]

    for label, value in stats:
        ws_resumen.cell(row=row, column=1, value=label).font = label_font
        ws_resumen.cell(row=row, column=2, value=value).font = value_font
        row += 1

    # Por tipo
    row += 1
    ws_resumen.cell(row=row, column=1, value="Por Tipo").font = subtitle_font
    row += 1
    tipos = Counter(t.get("tipo", "Otro") for t in tickets)
    for tipo, count in sorted(tipos.items()):
        ws_resumen.cell(row=row, column=1, value=tipo).font = label_font
        ws_resumen.cell(row=row, column=2, value=count).font = value_font
        row += 1

    # Por prioridad
    row += 1
    ws_resumen.cell(row=row, column=1, value="Por Prioridad").font = subtitle_font
    row += 1
    prioridades = Counter(t.get("prioridad", "Sin prioridad") for t in tickets)
    for prio, count in sorted(prioridades.items()):
        ws_resumen.cell(row=row, column=1, value=prio).font = label_font
        ws_resumen.cell(row=row, column=2, value=count).font = value_font
        row += 1

    # Por categoria
    row += 1
    ws_resumen.cell(row=row, column=1, value="Por Categoría").font = subtitle_font
    row += 1
    categorias = Counter(t.get("categoria", "Sin categoría") for t in tickets)
    for cat, count in sorted(categorias.items()):
        ws_resumen.cell(row=row, column=1, value=cat).font = label_font
        ws_resumen.cell(row=row, column=2, value=count).font = value_font
        row += 1

    ws_resumen.column_dimensions["A"].width = 25
    ws_resumen.column_dimensions["B"].width = 15

    # --- Hoja 2: Tickets ---
    ws_tickets = wb.create_sheet("Tickets")

    headers = [
        "ID", "Título", "Tipo", "Estado", "Fecha Apertura",
        "Fecha Cierre", "Prioridad", "Categoría", "Solicitante", "Solución"
    ]
    col_widths = [10, 40, 14, 14, 16, 16, 14, 18, 25, 50]
    header_style = create_header_style()

    for col, header in enumerate(headers, 1):
        cell = ws_tickets.cell(row=1, column=col, value=header)
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.alignment = header_style["alignment"]
        cell.border = header_style["border"]

    for col, width in enumerate(col_widths, 1):
        ws_tickets.column_dimensions[get_column_letter(col)].width = width

    for row_idx, ticket in enumerate(tickets, 2):
        values = [
            ticket.get("id", ""),
            ticket.get("titulo", ""),
            ticket.get("tipo", ""),
            ticket.get("estado", ""),
            ticket.get("fecha_apertura", ""),
            ticket.get("fecha_cierre", ""),
            ticket.get("prioridad", ""),
            ticket.get("categoria", ""),
            ticket.get("solicitante", ""),
            ticket.get("solucion", ""),
        ]
        style = create_cell_style(row_idx)
        for col, value in enumerate(values, 1):
            cell = ws_tickets.cell(row=row_idx, column=col, value=value)
            cell.fill = style["fill"]
            cell.border = style["border"]
            cell.alignment = style["alignment"]

    # Auto-filtro y freeze
    last_col = get_column_letter(len(headers))
    ws_tickets.auto_filter.ref = f"A1:{last_col}{len(tickets) + 1}"
    ws_tickets.freeze_panes = "A2"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Excel generado: {output_path} ({len(tickets)} tickets)", file=sys.stderr)


def main():
    if len(sys.argv) < 3:
        print("Uso: python3 scripts/glpi_to_excel.py <json_input> <xlsx_output>")
        sys.exit(1)

    json_input = sys.argv[1]
    xlsx_output = sys.argv[2]

    with open(json_input, "r", encoding="utf-8") as f:
        data = json.load(f)

    generate_excel(data, xlsx_output)


if __name__ == "__main__":
    main()
